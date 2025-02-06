from selenium import webdriver
import random
import openpyxl
import os

# Generate random name
first_names = ["John", "Jane", "Alice", "Bob", "Charlie", "Daisy", "Eve", "Frank"]
last_names = ["Smith", "Johnson", "Brown", "Williams", "Jones", "Miller", "Davis"]
name = random.choice(first_names) + " " + random.choice(last_names)

# Generate random email
domains = ["gmail.com", "yahoo.com", "outlook.com", "hotmail.com"]
username = name.lower().replace(" ", "") + str(random.randint(10, 99))
email = username + "@" + random.choice(domains)

# Generate random phone number
phone = "+1-" + str(random.randint(100, 999)) + "-" + str(random.randint(100, 999)) + "-" + str(random.randint(1000, 9999))

print("Name:", name)
print("Email:", email)
print("Phone:", phone)

# Define the Excel file path
file_path = "generated_data.xlsx"

# Check if the file exists
if os.path.exists(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Email", "Phone"])  # Add header row if file is new

# Append the generated data to the Excel file
sheet.append([name, email, phone])

# Save the file
workbook.save(file_path)

workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Get the last row number
last_row = sheet.max_row

# Get the "Name" and "Email" data from the last row (columns 1 and 2)
name = sheet.cell(row=last_row, column=1).value
email = sheet.cell(row=last_row, column=2).value

# Print the "Name" and "Email" from the last row
print("Latest Row Data:")
print("Name:", name)
print("Email:", email)