import time
from openpyxl import load_workbook
import pytest
import allure
from allure_commons.types import AttachmentType
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Load the Excel file
file_path = "test_data.xlsx"
wb = load_workbook(filename=file_path, data_only=True)
ws = wb.active  # Get the active sheet

# Define the specific stop value
stop_value = "Test Usman"
column_index = 1  # Second column (B)
filtered_values = []

# Extract values from the second column, from bottom to top
for row in reversed(range(1, ws.max_row + 1)):  # Start from last row to first row
    cell_value = ws.cell(row=row, column=column_index).value
    if cell_value is not None:
        filtered_values.append(cell_value)
        if cell_value == stop_value:
            break  # Stop when the specific value is found

filtered_values.reverse()  # Restore correct order

# Print or return the extracted values
print(filtered_values)

@pytest.fixture(scope="module")
def driver():
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.implicitly_wait(10)
    yield driver
    driver.quit()

def test_create_account_single_record(driver, config):
    driver.get("https://test.salesforce.com/")
    wait = WebDriverWait(driver, 20)

    # Perform login
    username = wait.until(EC.element_to_be_clickable((By.ID, "username")))
    username.send_keys("draftsf@draftdata.com.uat")
    password = wait.until(EC.element_to_be_clickable((By.ID, "password")))
    password.send_keys("Rolustech@99")
    login_button = wait.until(EC.element_to_be_clickable((By.ID, "Login")))
    login_button.click()
    driver.delete_all_cookies()

    # Move to account Tab and click on new button
    driver.get("https://dakotanetworks--uat.sandbox.lightning.force.com/lightning/o/Contact/list?filterName=All_Contacts")
    # button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@title='Select a List View: Accounts']//lightning-primitive-icon[@exportparts='icon']//*[name()='svg']")))
    # button.click()
    # button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='Search lists...']")))
    # button.send_keys("All Accounts")
    # button = wait.until(EC.element_to_be_clickable((By.XPATH, "//li[1]//a[1]//span[1]//mark[1]")))
    # button.click()

    # Search the result
    for value in filtered_values:
        print(f"Processing value: {value}")
        search_input = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@name='Contact-search-input']")))
        search_input.clear()
        search_input.send_keys(value + Keys.ENTER)
        driver.find_element(By.XPATH, "//button[@title='Refresh']//lightning-primitive-icon[@exportparts='icon']").click()
        time.sleep(3)
        try:
            edit_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@class='slds-icon_container slds-icon-utility-down']//span[1]")))
            edit_btn.click()
            delete_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "(//a[@title='Delete'])[1]")))
            delete_btn.click()
            delete_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[normalize-space()='Delete']")))
            delete_btn.click()
            time.sleep(3)
        except:
            print(f"Account {value} not found")
        time.sleep(3)



